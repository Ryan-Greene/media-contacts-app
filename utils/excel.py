import io
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS     = ["Outlet","Contact First","Contact Last","Title","Email","Phone",
               "Market Type","Market","Client(s)","Notes","Website"]
COL_WIDTHS  = [30,14,14,28,30,16,12,14,22,35,25]

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SECTION_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
BODY_FONT    = Font(name="Calibri", size=10)
WRAP         = Alignment(wrap_text=True, vertical="top")
THIN         = Side(style="thin", color="B8CCE4")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MAIN_TABS   = ["Print & Online","Broadcast (TV)","Radio","Newsletter","Podcast"]
TRADE_ORDER = ["Trade - Architecture","Trade - Classical Music","Trade - Disability",
               "Trade - Education","Trade - Health","Trade - Insurance",
               "Trade - Military","Trade - Philanthropy","Trade - General"]

def _style_header(ws):
    ws.row_dimensions[1].height = 20
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

def _section_row(ws, row_num, label):
    for col in range(1, len(HEADERS)+1):
        c = ws.cell(row=row_num, column=col)
        c.fill   = SECTION_FILL
        c.border = BORDER
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font      = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 16
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num,   end_column=len(HEADERS))

def _data_row(ws, row_num, c, alt=False):
    vals = [c.get("Outlet",""), c.get("Contact First",""), c.get("Contact Last",""),
            c.get("Title",""),  c.get("Email",""),         c.get("Phone",""),
            c.get("Market Type",""), c.get("Market",""),   c.get("Client(s)",""),
            c.get("Notes",""),  c.get("Website","")]
    fill = ALT_FILL if alt else None
    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=col, value=val or "")
        cell.font      = BODY_FONT
        cell.alignment = WRAP
        cell.border    = BORDER
        if fill: cell.fill = fill

def _write_flat(ws, records):
    _style_header(ws)
    for i, c in enumerate(sorted(records, key=lambda x: (x.get("Outlet") or "").lower())):
        _data_row(ws, i+2, c, alt=(i%2==0))

def _write_sectioned(ws, sections):
    _style_header(ws)
    row = 2
    for label, recs in sections:
        if not recs: continue
        _section_row(ws, row, label)
        row += 1
        for i, c in enumerate(sorted(recs, key=lambda x: (x.get("Outlet") or "").lower())):
            _data_row(ws, row, c, alt=(i%2==0))
            row += 1

def build_excel(contacts: list[dict], list_name: str = "Media List") -> bytes:
    wb = openpyxl.Workbook()

    # All Contacts
    ws_all = wb.active
    ws_all.title = "All Contacts"
    _write_flat(ws_all, contacts)

    # Main tabs
    for tab in MAIN_TABS:
        recs = [c for c in contacts if c.get("Media Type") == tab]
        if not recs: continue
        ws = wb.create_sheet(tab)
        _write_flat(ws, recs)

    # Trade Media with sub-sections
    trade_recs = [c for c in contacts if c.get("Media Type") == "Trade Media"]
    if trade_recs:
        ws_t = wb.create_sheet("Trade Media")
        sections = []
        for sub in TRADE_ORDER:
            sub_recs = [c for c in trade_recs if c.get("Trade Sub") == sub]
            if sub_recs: sections.append((sub, sub_recs))
        no_sub = [c for c in trade_recs if not c.get("Trade Sub")]
        if no_sub: sections.append(("Trade - General", no_sub))
        _write_sectioned(ws_t, sections)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
